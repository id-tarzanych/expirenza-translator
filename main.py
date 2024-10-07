import os
import sys
import time
import json
import getopt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
from openpyxl import Workbook

from openai import OpenAI
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY")
)

# Function to translate text using OpenAI API
def translate_text(text, target_language):
    if not text:
        return ""
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": f"You are a helpful assistant that translates Ukrainian to {target_language}. "
                               f"Provide only the translation without any additional text or explanation. "
                               f"If there is no direct translation or the term is the same in {target_language}, return the original text."
                },
                {
                    "role": "user",
                    "content": f"Translate the following Ukrainian text to {target_language}:\n\n{text}"
                }
            ],
            max_tokens=500,
            temperature=0.1,
        )
        translation = response.choices[0].message.content.strip()
        return translation
    except Exception as e:
        print(f"Translation error: {e}")
        return ""


def main(argv):
    # Default values
    output_file = None
    restaurant_uuid = None
    target_language = "English"

    # Parse command-line arguments
    try:
        opts, args = getopt.getopt(argv, "hu:l:", ["uuid=", "language="])
    except getopt.GetoptError:
        print('Usage: script_name.py output_file.xlsx -u <restaurant_uuid> [-l <language>]')
        sys.exit(2)

    # First command-line argument is the output file
    if len(args) > 0:
        output_file = args[0]
    else:
        print('Error: Output file path is required as the first argument.')
        print('Usage: script_name.py output_file.xlsx -u <restaurant_uuid> [-l <language>]')
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print('Usage: script_name.py output_file.xlsx -u <restaurant_uuid> [-l <language>]')
            sys.exit()
        elif opt in ("-u", "--uuid"):
            restaurant_uuid = arg
        elif opt in ("-l", "--language"):
            target_language = arg

    if not restaurant_uuid:
        print('Error: Restaurant UUID is required. Use -u <restaurant_uuid> to specify it.')
        sys.exit(2)

    # Set up OpenAI API key
    if not os.getenv("OPENAI_API_KEY"):
        print("Error: OpenAI API key not found. Set the OPENAI_API_KEY environment variable.")
        sys.exit(1)

    # Initialize Excel workbook
    wb = Workbook()

    # Create sheets
    ws_categories = wb.create_sheet("Categories and Subcategories")
    ws_menu_items = wb.create_sheet("Menu Items")
    ws_restaurant_details = wb.create_sheet("Restaurant Details")

    # Remove default sheet
    del wb['Sheet']

    # Set up headers
    ws_categories.append(["Ukrainian text", f"{target_language} translation"])
    ws_menu_items.append(["Category", "Subcategory", "Item Title", "Type", "Ukrainian text", f"{target_language} translation"])
    ws_restaurant_details.append(["Ukrainian text", f"{target_language} translation"])

    # Adjust column widths for better readability
    for ws in [ws_categories, ws_menu_items, ws_restaurant_details]:
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50
        if ws == ws_menu_items:
            ws.column_dimensions['C'].width = 50  # Item Title
            ws.column_dimensions['D'].width = 20  # Type
            ws.column_dimensions['E'].width = 50  # Ukrainian text
            ws.column_dimensions['F'].width = 50  # Translation

    # Set up Selenium WebDriver
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run headless Chrome
    driver = webdriver.Chrome(options=chrome_options)

    try:
        # 1. Open cafe page on Expirenza
        url = f"https://expz.menu/{restaurant_uuid}"
        print(f"Opening URL: {url}")
        driver.get(url)
        time.sleep(5)  # Wait for the page to load

        # 2. From the initial page crawl Subtitle text and About restaurant text
        # 2.1. Subtitle
        try:
            subtitle_element = driver.find_element(By.CSS_SELECTOR, "div.subtitle")
            subtitle_text = subtitle_element.text.strip()
            print(f"Crawled subtitle text: {subtitle_text}")
            subtitle_translation = translate_text(subtitle_text, target_language)
            ws_restaurant_details.append([subtitle_text, subtitle_translation])
        except Exception as e:
            print(f"Error retrieving subtitle: {e}")

        # 2.2. About restaurant text
        try:
            about_element = driver.find_element(By.CSS_SELECTOR, "div.decorate-supported-text")
            about_text = about_element.text.strip()
            print(f"Crawled about text: {about_text}")
            about_translation = translate_text(about_text, target_language)
            ws_restaurant_details.append([about_text, about_translation])
        except Exception as e:
            print(f"Error retrieving about restaurant text: {e}")

        # 3. Click on the first menu item to proceed to the menu page
        try:
            first_menu_item = driver.find_element(By.CSS_SELECTOR, "a.main-menu-item")
            print(f"Clicking on first menu item: {first_menu_item.text.strip()}")
            driver.execute_script("arguments[0].click();", first_menu_item)
            time.sleep(3)  # Wait for navigation
        except Exception as e:
            print(f"Error clicking on first menu item: {e}")
            return

        # Get the number of main categories
        num_categories = len(driver.find_elements(By.CSS_SELECTOR, "li.side-menu__category"))
        print(f"Found {num_categories} main categories.")

        for i in range(1, num_categories + 1):
            # Build the category CSS selector
            category_selector = f"li.side-menu__category:nth-of-type({i})"

            # 4.1 Crawl main category name
            try:
                category_title_element = driver.find_element(By.CSS_SELECTOR, f"{category_selector} a.side-menu__category--link div")
                category_title = category_title_element.text.strip()
                print(f"Crawled category title: {category_title}")
                category_translation = translate_text(category_title, target_language)
                ws_categories.append([category_title, category_translation])
            except Exception as e:
                print(f"Error retrieving category title: {e}")
                continue

            # 4.1 Click on item to open category page
            try:
                category_link = driver.find_element(By.CSS_SELECTOR, f"{category_selector} a.side-menu__category--link")
                print(f"Clicking on category: {category_title}")
                driver.execute_script("arguments[0].click();", category_link)
                time.sleep(0.1)  # Wait for page to load
            except Exception as e:
                print(f"Error clicking on category: {e}")
                continue

            # 4.2 From the dish list crawl each subcategory list block
            try:
                # Get number of dish lists
                dish_lists = driver.find_elements(By.CSS_SELECTOR, "div.dish-list")
                num_dish_lists = len(dish_lists)
                print(f"Found {num_dish_lists} dish lists in category '{category_title}'.")
            except Exception as e:
                print(f"Error retrieving dish lists: {e}")
                continue

            for j in range(1, num_dish_lists + 1):
                dish_list_selector = f"div.dish-list:nth-of-type({j})"
                # 4.2.1 Subcategory title
                try:
                    subcategory_title_element = driver.find_element(By.CSS_SELECTOR, f"{dish_list_selector} h2.dish-list--title")
                    subcategory_title = subcategory_title_element.text.strip()
                    print(f"Crawled subcategory title: {subcategory_title}")
                    subcategory_translation = translate_text(subcategory_title, target_language)
                    ws_categories.append([subcategory_title, subcategory_translation])
                except Exception as e:
                    print(f"Error retrieving subcategory title: {e}")
                    continue

                # 4.2.2 Dishes in subcategory
                try:
                    menu_items = driver.find_elements(By.CSS_SELECTOR, f"{dish_list_selector} div.menu-list-item")
                    num_menu_items = len(menu_items)
                    print(f"Found {num_menu_items} menu items in subcategory '{subcategory_title}'.")
                except Exception as e:
                    print(f"Error retrieving menu items: {e}")
                    continue

                for k in range(1, num_menu_items + 1):
                    menu_item_selector = f"{dish_list_selector} div.menu-list-item:nth-of-type({k})"
                    # 4.2.2.1 Dish item title
                    try:
                        item_title_element = driver.find_element(By.CSS_SELECTOR, f"{menu_item_selector} h4.item-title")
                        item_title = item_title_element.text.strip()
                        print(f"Crawled item title: {item_title}")
                        item_title_translation = translate_text(item_title, target_language)
                        # Append the item title row
                        ws_menu_items.append([category_title, subcategory_title, item_title, "Title", item_title, item_title_translation])
                    except Exception as e:
                        print(f"Error retrieving item title: {e}")
                        continue

                    # 4.2.2.2 Dish item amount (optional)
                    try:
                        item_amount_element = driver.find_element(By.CSS_SELECTOR, f"{menu_item_selector} div.item-amount")
                        item_amount = item_amount_element.text.strip()
                        print(f"Crawled item amount: {item_amount}")
                        item_amount_translation = translate_text(item_amount, target_language)
                        ws_menu_items.append([category_title, subcategory_title, item_title, "Amount", item_amount, item_amount_translation])
                    except Exception as e:
                        print(f"Item amount not found for '{item_title}'. Continuing...")

                    # 4.2.2.3 Click on dish item title, get description from popup
                    try:
                        item_title_link = driver.find_element(By.CSS_SELECTOR, f"{menu_item_selector} h4.item-title")
                        print(f"Clicking on item to get description: {item_title}")
                        driver.execute_script("arguments[0].click();", item_title_link)
                        time.sleep(0.1)  # Wait for popup

                        # Get description from popup (modified)
                        try:
                            description_container = driver.find_element(By.CSS_SELECTOR, "div.dish--description")
                            description_elements = description_container.find_elements(By.CSS_SELECTOR, "p")
                            if description_elements:
                                description_texts = [elem.text.strip() for elem in description_elements]
                                description_text = "\n".join(description_texts)
                            else:
                                # If no <p> elements, get the text of the container
                                description_text = description_container.text.strip()
                            print(f"Crawled item description: {description_text}")
                            description_translation = translate_text(description_text, target_language)
                            ws_menu_items.append([category_title, subcategory_title, item_title, "Description", description_text, description_translation])
                        except Exception as e:
                            print(f"Item description not found for '{item_title}'. Continuing...")

                        # Always attempt to crawl allergens
                        try:
                            allergens_element = driver.find_element(By.CSS_SELECTOR, "div.allergens--list")
                            allergens_text = allergens_element.text.strip()
                            print(f"Crawled allergens: {allergens_text}")
                            allergens_translation = translate_text(allergens_text, target_language)
                            ws_menu_items.append([category_title, subcategory_title, item_title, "Allergens", allergens_text, allergens_translation])
                        except Exception as e:
                            print(f"No allergens found for '{item_title}'. Continuing...")

                        # Close popup
                        try:
                            close_button = driver.find_element(By.CSS_SELECTOR, "div.close-btn")
                            driver.execute_script("arguments[0].click();", close_button)
                            time.sleep(0.1)
                        except Exception as e:
                            print(f"Error closing popup: {e}")
                            from selenium.webdriver.common.keys import Keys
                            driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                            time.sleep(0.1)
                    except Exception as e:
                        print(f"Error clicking on item title: {e}")
                        continue

        # Save the workbook
        wb.save(output_file)
        print(f"Excel file '{output_file}' has been created successfully.")

    finally:
        # Close the browser
        driver.quit()

if __name__ == "__main__":
    main(sys.argv[1:])
